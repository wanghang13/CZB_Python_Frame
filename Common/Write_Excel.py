# *****************写数据*****************
# 先复制一份测试用例
# 在写入测试数据
# 作者：杭仔
# ****************************************
# -*- coding:UTF-8 -*-
import sys,os
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from openpyxl import load_workbook
from openpyxl.styles import colors
from openpyxl.styles import Font
import openpyxl
import os
import shutil
from Common import Log


def copy_file(file_path1, file_path2):
    """复制文件"""
    result = os.path.exists(file_path1)
    if result is True:
        shutil.copyfile(file_path1, file_path2)
    else:
        Log.TestLog().error("{}复制文件不存在".format(file_path1))


def copy_excel(excel_path1, excel_path2):
    """
    复制excel，把excelpath1数据复制到excelpath2
    :param excel_path1: 要复制的excel路径（到文件）
    :param excel_path2: 复制后的excel路径（到文件）
    :return:
    """
    result = os.path.exists(excel_path1)
    if result is True:
        # 获取测试用例表格
        wb1 = openpyxl.load_workbook(excel_path1)
        sheets1 = wb1.sheetnames
        # 复制excel
        wb2 = openpyxl.Workbook()
        for sheet_name in sheets1:
            wb2.create_sheet(sheet_name)
        # 删除初始的Sheet表
        wb2.remove_sheet(wb2.get_sheet_by_name("Sheet"))
        # 保存复制的excel
        wb2.save(excel_path2)
        # 获取复制的表格
        wb2 = openpyxl.load_workbook(excel_path2)
        sheets2 = wb2.sheetnames
        # 复制测试用例表格数据
        for j in range(len(sheets1)):
            sheet1 = wb1[sheets1[j]]
            sheet2 = wb2[sheets2[j]]
            max_row = sheet1.max_row  # 最大行数
            max_column = sheet1.max_column  # 最大列数
            for m in list(range(1, max_row + 1)):
                for n in list(range(97, 97 + max_column)):  # chr(97)='a'
                    n = chr(n)  # ASCII字符(A,B...)
                    i = '%s%d' % (n, m)  # 单元格编号(A1,A2...)
                    cell1 = sheet1[i].value  # 获取data单元格数据
                    sheet2[i].value = cell1  # 赋值到test单元格
        wb2.save(excel_path2)  # 保存数据
        wb1.close()  # 关闭excel
        wb2.close()
    else:
        Log.TestLog().error("{}复制文件不存在".format(excel_path1))


class WriteExcel(object):
    """修改excel数据"""

    def __init__(self, filename, sheet_name="Sheet_basic"):
        """
        修改excel数据
        :param filename: 文件路径（到文件）
        :param sheet_name: sheet名称，默认Sheet_basic
        """
        self.filename = filename
        self.wb = load_workbook(self.filename)
        # self.ws = self.wb.active  # 激活sheet
        self.sheet = self.wb[sheet_name]

    def write(self, row_n, col_n, value, colour=None):
        """
        写入数据，如(2,3，"hello"),第二行第三列写入数据"hello"
        :param row_n: 行号
        :param col_n: 列号
        :param value: 内容
        :param colour: 颜色，默认为None，其他为红色
        :return:
        """
        if colour == "red":
            colour = Font(color=colors.RED)  # color="FFBB00"，颜色编码也可以设定颜色
            self.sheet.cell(row_n, col_n).font = colour
            self.sheet.cell(row_n, col_n).value = value
        elif colour == "green":
            colour = Font(color=colors.BLUE)  # color="FFBB00"，颜色编码也可以设定颜色
            self.sheet.cell(row_n, col_n).font = colour
            self.sheet.cell(row_n, col_n).value = value
        else:
            self.sheet.cell(row_n, col_n).value = value
        self.wb.save(self.filename)
